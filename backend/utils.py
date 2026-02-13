#utils
import pandas as pd
import io
import re

def generate_excel(data):
    """Generate a visually styled Excel file from a list/dict `data`.

    Improvements:
    - Bold header with green fill and white text
    - Freeze header row
    - Auto column widths (supports >26 columns)
    - Numeric formatting for numeric columns
    - Add a styled Excel table with banded rows and filters
    """
    if not data:
        df = pd.DataFrame([{"Alert": "No data found"}])
    else:
        df = pd.DataFrame(data)

    # Attempt to normalize and coerce numeric-like strings into numbers so Excel shows them properly
    def parse_numeric(val):
        try:
            if val is None:
                return None
            s = str(val).strip()
            if s == '':
                return None
            # Common dash or em-dash used to indicate empty
            if s in ['–', '—', '—', '-']:
                return None
            # Parentheses denote negative numbers e.g. (1,234)
            negative = False
            if s.startswith('(') and s.endswith(')'):
                negative = True
                s = s[1:-1]
            # Remove commas and non-numeric chars except dot and minus
            s2 = re.sub(r'[^0-9\.\-]', '', s)
            if s2 == '' or s2 == '.' or s2 == '-':
                return None
            # If multiple dots, keep first and remove others
            if s2.count('.') > 1:
                parts = s2.split('.')
                s2 = parts[0] + '.' + ''.join(parts[1:])
            num = float(s2)
            if negative:
                num = -num
            # Return int when possible
            if abs(num - int(num)) < 1e-9:
                return int(num)
            return num
        except Exception:
            return None

    # Apply numeric parsing column-wise and decide which columns are numeric
    numeric_cols = []
    if not df.empty:
        for col in df.columns:
            parsed = df[col].astype(object).apply(parse_numeric)
            # if more than 30% of non-empty values parsed as numbers, treat column as numeric
            non_empty = df[col].astype(str).replace('nan', '').replace('None', '').map(lambda x: x.strip() if isinstance(x, str) else '').map(bool)
            non_empty_count = non_empty.sum()
            parsed_count = parsed.notnull().sum()
            if non_empty_count == 0:
                is_numeric = False
            else:
                is_numeric = (parsed_count / max(1, non_empty_count)) >= 0.3

            if is_numeric and parsed_count > 0:
                numeric_cols.append(col)
                df[col] = parsed

        # Replace remaining empty strings/NaN with None for later dash rendering
        df = df.where(pd.notnull(df), None)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Full_Report')

        # Work with the underlying openpyxl worksheet
        workbook = writer.book
        worksheet = writer.sheets['Full_Report']

        try:
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except Exception:
            # If openpyxl missing, return raw bytes (pandas will still have written)
            return output.getvalue()

        # Header style
        header_font = Font(bold=True, color='FFFFFFFF')
        header_fill = PatternFill(fill_type='solid', start_color='FF2E7D32', end_color='FF2E7D32')
        header_align = Alignment(horizontal='center', vertical='center')
        thin = Side(border_style='thin', color='FFDDDDDD')
        header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Apply header styles (first row)
        for col_idx, col in enumerate(df.columns, 1):
            # Create a cleaned, human-friendly header label
            cleaned = re.sub(r'_+', ' ', str(col)).strip()
            cleaned = re.sub(r'\s+', ' ', cleaned)
            cleaned_label = cleaned.title()
            if len(cleaned_label) > 36:
                cleaned_label = cleaned_label[:33] + '...'

            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = cleaned_label
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = header_border

        # Freeze header row
        worksheet.freeze_panes = worksheet['A2']

        # Determine column widths and apply number formats
        for i, col in enumerate(df.columns):
            col_idx = i + 1
            col_letter = get_column_letter(col_idx)

            # Compute width based on header and cell contents
            max_length = len(str(col))
            if not df.empty:
                # Convert values to string for length check
                col_series = df[col].astype(str)
                max_length = max(max_length, col_series.map(len).max())

            adjusted_width = min((max_length + 2), 60)
            worksheet.column_dimensions[col_letter].width = adjusted_width

            # Apply numeric formatting if column is numeric
            try:
                # Prefer using our detected numeric columns list
                if col in numeric_cols:
                    # pick a sample non-null value to decide int vs float
                    sample = None
                    for v in df[col]:
                        if v is not None:
                            sample = v
                            break
                    if isinstance(sample, int):
                        fmt = '#,##0;(#,##0)'
                    else:
                        fmt = '#,##0.00;(#,##0.00)'

                    for row in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row, column=col_idx)
                        # Apply format only to numeric cell values
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = fmt
            except Exception:
                pass

        # Add an Excel table for better visuals (filters, banded rows)
        try:
            last_col = get_column_letter(len(df.columns) if len(df.columns) > 0 else 1)
            last_row = len(df) + 1
            table_ref = f'A1:{last_col}{last_row}'
            table = Table(displayName='FinancialReport', ref=table_ref)
            style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            worksheet.add_table(table)
        except Exception:
            # If table creation fails, ignore and continue
            pass

        # Add light borders to data cells for readability
        thin = Side(border_style='thin', color='FFEEEEEE')
        data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                # keep existing number formats and alignment but add border
                try:
                    # Replace empty / None with a dash for visual clarity
                    if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ''):
                        cell.value = '-'
                        cell.alignment = Alignment(horizontal='center')
                    cell.border = data_border
                except Exception:
                    pass

    return output.getvalue()